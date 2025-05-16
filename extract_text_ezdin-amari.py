import os
import logging
import pytesseract
import fitz  # PyMuPDF
import docx
import openpyxl
from openpyxl import load_workbook
from pdf2image import convert_from_path
from PIL import Image
from tqdm import tqdm

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

def pdf_to_image(pdf_path):
    """
    Convert PDF pages to images for OCR if needed.
    Requires pdf2image and Pillow libraries.
    """
    try:
        images = convert_from_path(pdf_path)
        return images
    except Exception as e:
        logging.error(f"Failed to convert PDF to image: {str(e)}")
        return []

def extract_text(upload_folder):
    """
    Extract text from files in the upload_folder.
    Supported formats: .pdf, .docx, .xls, .xlsx.
    Uses OCR for PDFs without extractable text.
    """
    if not os.path.exists(upload_folder):
        logging.error(f"Upload folder '{upload_folder}' does not exist.")
        return {}

    extracted_text = {}
    
    for filename in os.listdir(upload_folder):
        file_path = os.path.join(upload_folder, filename)
        if not filename.lower().endswith(('.pdf', '.docx', '.xls', '.xlsx')):
            logging.info(f"Skipping unsupported file: {filename}")
            continue

        try:
            if filename.lower().endswith('.pdf'):
                text = ""
                try:
                    with fitz.open(file_path) as pdf:
                        # First pass: Try native text extraction (faster, works for digital PDFs)
                        for page in pdf:
                            page_text = page.get_text()
                            if page_text:
                                text += page_text + "\n"
                        
                        # Second pass: If no text found, PDF might be scanned/image-based
                        if not text.strip():
                            logger.info(f"No text directly extracted from {file_path}. Attempting OCR.")
                            full_ocr_text = ""
                            
                            # Show progress for OCR which can be time-consuming
                            for i, page in enumerate(tqdm(pdf, desc="OCR Processing")):
                                try:
                                    # Convert page to image at 300dpi for OCR
                                    pix = page.get_pixmap(matrix=fitz.Matrix(300/72, 300/72))
                                    img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
                                    page_ocr_text = pytesseract.image_to_string(img, lang='eng')
                                    if page_ocr_text:
                                        full_ocr_text += page_ocr_text + "\n"
                                except Exception as ocr_page_error:
                                    logger.error(f"Error during OCR for page {i+1} of {file_path}: {ocr_page_error}")
                            text = full_ocr_text
                except Exception as e:
                    logger.error(f"Error processing PDF {file_path}: {str(e)}")
                    text = ""
                extracted_text[filename] = text
            elif filename.lower().endswith('.docx'):
                doc = docx.Document(file_path)
                text = "\n".join(paragraph.text for paragraph in doc.paragraphs)
                extracted_text[filename] = text
            elif filename.lower().endswith(('.xls', '.xlsx')):
                text = ""
                workbook = load_workbook(filename=file_path, read_only=True, data_only=True)
                for sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]
                    text += f"\n--- Sheet: {sheet_name} ---\n"
                    for row in sheet.iter_rows():
                        row_text = ""
                        for cell in row:
                            if cell.value is not None:
                                row_text += str(cell.value) + "\t"
                        if row_text:
                            text += row_text + "\n"
                extracted_text[filename] = text
            logging.info(f"Successfully extracted text from {filename}")
        except Exception as e:
            logging.error(f"Error processing {filename}: {str(e)}")
            continue
    
    return extracted_text

if __name__ == "__main__":
    upload_folder = "uploads"
    result = extract_text(upload_folder)
    # Save extracted text to files for verification
    output_dir = "extracted_output"
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
    for filename, text in result.items():
        output_path = os.path.join(output_dir, f"extracted_{filename}.txt")
        with open(output_path, "w", encoding="utf-8") as f:
            f.write(text)
        logging.info(f"Saved extracted text to {output_path}")
